from __future__ import annotations

import argparse
import copy
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

BASE_YEAR = 2024
END_YEAR = 2060
YEARS = list(range(BASE_YEAR, END_YEAR + 1))
SUMMARY_PATH = Path("version3/水泥行业模型校准说明.md")
WORKBOOK_PATH = Path("version3/水泥行业减碳基准路径模型V3.xlsx")


@dataclass(frozen=True)
class ScenarioConfig:
    key: str
    label: str
    sheet_name: str
    intensity_anchors: Dict[int, float]
    production_anchors: Dict[int, float]
    share_anchors: Dict[int, List[int]]
    source_text: str
    scenario_text: str


SCENARIOS: Dict[str, ScenarioConfig] = {
    "conservative": ScenarioConfig(
        key="conservative",
        label="保守情景",
        sheet_name="保守情景（政策承诺）",
        intensity_anchors={2024: 0.590, 2030: 0.571, 2035: 0.540, 2040: 0.500, 2050: 0.340, 2060: 0.095},
        production_anchors={2024: 18.3, 2030: 14.2, 2035: 13.8, 2040: 12.8, 2050: 7.2, 2060: 6.0},
        share_anchors={
            2024: [100, 0, 0, 0, 0, 0],
            2030: [45, 25, 15, 8, 0, 7],
            2035: [40, 24, 18, 8, 3, 7],
            2040: [34, 22, 18, 10, 8, 8],
            2050: [28, 18, 20, 12, 14, 8],
            2060: [22, 14, 18, 16, 20, 10],
        },
        source_text="数据来源：CBMA《中国水泥行业碳中和路径研究》(2023) + 政策延续情景再校准（2030/2035关键节点下调）",
        scenario_text=(
            "保守情景以政策承诺与行业现实约束为主，产量下降相对平缓，2030年前主要依赖需求收缩、能效改造和少量替代燃料，"
            "CCUS在2035年前仅保持示范部署。"
        ),
    ),
    "moderate": ScenarioConfig(
        key="moderate",
        label="适度情景",
        sheet_name="适度情景（2°C路径）",
        intensity_anchors={2024: 0.590, 2030: 0.547, 2035: 0.496, 2040: 0.420, 2050: 0.230, 2060: 0.085},
        production_anchors={2024: 18.3, 2030: 13.4, 2035: 12.9, 2040: 11.7, 2050: 6.5, 2060: 5.5},
        share_anchors={
            2024: [100, 0, 0, 0, 0, 0],
            2030: [45, 22, 16, 8, 0, 9],
            2035: [38, 21, 19, 10, 5, 7],
            2040: [32, 19, 20, 12, 10, 7],
            2050: [24, 15, 18, 15, 20, 8],
            2060: [20, 10, 16, 14, 28, 12],
        },
        source_text="数据来源：IFS-CGE模型2°C路径 + 行业运行数据校准 + 2030/2035关键节点约束修正",
        scenario_text=(
            "适度情景作为推荐基准，维持中等产量下降速度，2030年前稳步推进能效改造、熟料替代和替代燃料，"
            "2030年后CCUS逐步商业化放量。"
        ),
    ),
    "aggressive": ScenarioConfig(
        key="aggressive",
        label="积极情景",
        sheet_name="积极情景（NZE路径）",
        intensity_anchors={2024: 0.590, 2030: 0.521, 2035: 0.454, 2040: 0.350, 2050: 0.180, 2060: 0.050},
        production_anchors={2024: 18.3, 2030: 12.8, 2035: 11.8, 2040: 10.5, 2050: 5.8, 2060: 4.8},
        share_anchors={
            2024: [100, 0, 0, 0, 0, 0],
            2030: [36, 24, 22, 10, 2, 6],
            2035: [30, 20, 22, 13, 9, 6],
            2040: [22, 16, 20, 14, 22, 6],
            2050: [16, 10, 16, 16, 34, 8],
            2060: [12, 8, 12, 16, 42, 10],
        },
        source_text="数据来源：IEA NZE情景 + 中国水泥行业情境化校准（放缓2030前抢跑节奏、强化2030后提速）",
        scenario_text=(
            "积极情景在宏观需求更快回落的同时，2030年前显著推进能效、替代燃料与低碳熟料，2030-2035年成为CCUS和深度脱碳加速窗口。"
        ),
    ),
}

TECH_LABELS = [
    "产能优化/产量下降",
    "极致能效提升",
    "替代燃料(TSR)",
    "熟料替代/低碳熟料",
    "CCUS碳捕集",
    "其他(碳汇/绿电等)",
]
KEY_YEARS = [2024, 2030, 2035, 2040, 2050, 2060]


def linear_interpolate(anchor_map: Dict[int, float], years: Iterable[int], digits: int) -> Dict[int, float]:
    anchors = sorted(anchor_map.items())
    results: Dict[int, float] = {}
    for year in years:
        if year in anchor_map:
            results[year] = round(anchor_map[year], digits)
            continue
        lower = max(a for a, _ in anchors if a < year)
        upper = min(a for a, _ in anchors if a > year)
        low_val = anchor_map[lower]
        up_val = anchor_map[upper]
        ratio = (year - lower) / (upper - lower)
        value = low_val + (up_val - low_val) * ratio
        results[year] = round(value, digits)
    return results


def interpolate_share_vectors(anchor_map: Dict[int, List[int]], years: Iterable[int]) -> Dict[int, List[float]]:
    anchors = sorted(anchor_map.items())
    results: Dict[int, List[float]] = {}
    for year in years:
        if year in anchor_map:
            results[year] = [float(v) for v in anchor_map[year]]
            continue
        lower = max(a for a, _ in anchors if a < year)
        upper = min(a for a, _ in anchors if a > year)
        low_vec = anchor_map[lower]
        up_vec = anchor_map[upper]
        ratio = (year - lower) / (upper - lower)
        vec = [round(low + (up - low) * ratio, 1) for low, up in zip(low_vec, up_vec)]
        total = round(sum(vec), 1)
        if total != 100.0:
            vec[-1] = round(vec[-1] + (100.0 - total), 1)
        results[year] = vec
    return results


def format_drop(base: float, value: float) -> float:
    return round((value - base) / base, 3)


def copy_row_style(ws: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        dst._style = copy.copy(src._style)
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy.copy(src.protection)


def build_paths(config: ScenarioConfig) -> Dict[str, Dict[int, float] | Dict[int, List[float]]]:
    intensity = linear_interpolate(config.intensity_anchors, YEARS, 3)
    production = linear_interpolate(config.production_anchors, YEARS, 1)
    emissions = {year: round(production[year] * intensity[year], 2) for year in YEARS}
    base = intensity[BASE_YEAR]
    drops = {year: format_drop(base, intensity[year]) for year in YEARS}
    shares = interpolate_share_vectors(config.share_anchors, YEARS)
    reductions = {year: round(base - intensity[year], 3) for year in YEARS}
    absolute_contribs: Dict[int, List[float]] = {}
    for year in YEARS:
        total = reductions[year]
        if total == 0:
            absolute_contribs[year] = [0.0] * len(TECH_LABELS)
            continue
        contribs = [round(total * share / 100.0, 4) for share in shares[year]]
        rounding_gap = round(total - round(sum(contribs), 4), 4)
        contribs[-1] = round(contribs[-1] + rounding_gap, 4)
        absolute_contribs[year] = contribs
    return {
        "intensity": intensity,
        "production": production,
        "emissions": emissions,
        "drops": drops,
        "shares": shares,
        "reductions": reductions,
        "absolute_contribs": absolute_contribs,
    }


def update_path_sheet(ws: Worksheet, config: ScenarioConfig, paths: Dict[str, Dict[int, float] | Dict[int, List[float]]]) -> None:
    intensity = paths["intensity"]
    production = paths["production"]
    emissions = paths["emissions"]
    drops = paths["drops"]
    shares = paths["shares"]
    absolute_contribs = paths["absolute_contribs"]

    ws["A2"] = config.source_text
    for idx, year in enumerate(YEARS, start=46):
        ws[f"A{idx}"] = year
        ws[f"B{idx}"] = production[year]
        ws[f"C{idx}"] = emissions[year]
        ws[f"D{idx}"] = intensity[year]
        ws[f"E{idx}"] = drops[year]

    for col_offset, year in enumerate(KEY_YEARS, start=2):
        share_vector = shares[year]
        for row_offset, share in enumerate(share_vector, start=87):
            ws.cell(row_offset, col_offset).value = f"{int(round(share))}%"
        ws.cell(93, col_offset).value = "100%"

    for idx, year in enumerate(YEARS, start=97):
        ws[f"A{idx}"] = year
        contribs = absolute_contribs[year]
        for col, value in zip("BCDEFG", contribs):
            ws[f"{col}{idx}"] = round(value, 4)
        ws[f"H{idx}"] = round(sum(contribs), 4)

    ws["A136"] = (
        f"  2024年：碳强度={intensity[2024]:.3f} tCO2/t，较2024下降0.0%，CO2排放={emissions[2024]:.2f}亿吨"
        f" /  2030年：碳强度={intensity[2030]:.3f} tCO2/t，较2024下降{abs(drops[2030]) * 100:.1f}%，CO2排放={emissions[2030]:.2f}亿吨"
        f" /  2035年：碳强度={intensity[2035]:.3f} tCO2/t，较2024下降{abs(drops[2035]) * 100:.1f}%，CO2排放={emissions[2035]:.2f}亿吨"
        f" /  2040年：碳强度={intensity[2040]:.3f} tCO2/t，较2024下降{abs(drops[2040]) * 100:.1f}%，CO2排放={emissions[2040]:.2f}亿吨"
        f" /  2050年：碳强度={intensity[2050]:.3f} tCO2/t，较2024下降{abs(drops[2050]) * 100:.1f}%，CO2排放={emissions[2050]:.2f}亿吨"
        f" /  2060年：碳强度={intensity[2060]:.3f} tCO2/t，较2024下降{abs(drops[2060]) * 100:.1f}%，CO2排放={emissions[2060]:.2f}亿吨"
    )


def update_overview_sheet(ws: Worksheet, scenario_paths: Dict[str, Dict[str, Dict[int, float] | Dict[int, List[float]]]]) -> None:
    ws["A8"] = (
        "① 适度情景（2°C路径）：★推荐基准，采用‘自上而下’的2°C温控约束作为宏观锚，并结合行业实际运行数据对2030/2035关键节点进行校准；"
        "② 积极情景（NZE/1.5°C路径）：参考IEA NZE深度脱碳方向，但下调2030年前抢跑节奏、强化2030年后加速部署；"
        "③ 保守情景（政策承诺）：参考CBMA研究与现行政策执行节奏，仅在2035年前维持渐进式技术扩散。"
    )
    ws["A14"] = (
        "自上而下路径（宏观锚定）：三情景分别对应差异化水泥产量路径与碳预算约束，其中保守/适度/积极情景的2030年产量分别约为14.2/13.4/12.8亿吨，"
        "2035年分别约为13.8/12.9/11.8亿吨。 /自下而上路径（技术推演）：围绕能效提升、替代燃料、熟料替代、绿电与CCUS逐年分解碳强度下降。"
        " /耦合机制：以宏观产量和碳预算约束上边界，再以技术渗透节奏约束实现路径，保证2030/2035节点与2060终局同时可解释。"
    )
    ws["A20"] = (
        "① 宏观经济参数：北大国发院基准情景（GDP增速、人口、城镇化率） /② 产量路径：基于IFS-CGE行业产出中枢，并按保守/适度/积极三情景再校准"
        " /③ 碳强度校准：结合CBMA、IEA NZE、行业低碳技术渗透逻辑，重点满足2030与2035目标节点 /④ 技术路径分解：参考水泥行业低碳技术研究并按情景修正。"
    )
    ws["A23"] = (
        f"2030：水泥碳强度下降{abs(scenario_paths['conservative']['drops'][2030]) * 100:.1f}%–{abs(scenario_paths['aggressive']['drops'][2030]) * 100:.1f}%（保守-积极），"
        f"三情景产量分别约为{scenario_paths['conservative']['production'][2030]:.1f}/{scenario_paths['moderate']['production'][2030]:.1f}/{scenario_paths['aggressive']['production'][2030]:.1f}亿吨"
        f" /2035：碳强度下降{abs(scenario_paths['conservative']['drops'][2035]) * 100:.1f}%–{abs(scenario_paths['aggressive']['drops'][2035]) * 100:.1f}%，"
        "2030年前减排以需求收缩、能效和替代燃料为主，2030年后CCUS贡献逐步放大"
        f" /2040：碳强度下降{abs(scenario_paths['conservative']['drops'][2040]) * 100:.1f}%–{abs(scenario_paths['aggressive']['drops'][2040]) * 100:.1f}%"
        f" /2050：碳强度下降{abs(scenario_paths['conservative']['drops'][2050]) * 100:.1f}%–{abs(scenario_paths['aggressive']['drops'][2050]) * 100:.1f}%"
        f" /2060：碳强度下降{abs(scenario_paths['conservative']['drops'][2060]) * 100:.1f}%–{abs(scenario_paths['aggressive']['drops'][2060]) * 100:.1f}%，实现近零排放/碳中和导向。"
    )


def update_comparison_sheet(ws: Worksheet, scenario_paths: Dict[str, Dict[str, Dict[int, float] | Dict[int, List[float]]]]) -> None:
    cons = scenario_paths["conservative"]
    mod = scenario_paths["moderate"]
    agg = scenario_paths["aggressive"]

    for idx, year in enumerate(YEARS, start=6):
        ws[f"A{idx}"] = year
        ws[f"B{idx}"] = cons["intensity"][year]
        ws[f"C{idx}"] = cons["drops"][year]
        ws[f"D{idx}"] = mod["intensity"][year]
        ws[f"E{idx}"] = mod["drops"][year]
        ws[f"F{idx}"] = agg["intensity"][year]
        ws[f"G{idx}"] = agg["drops"][year]
        ws[f"H{idx}"] = round(mod["intensity"][year] - cons["intensity"][year], 3)
        ws[f"I{idx}"] = round(agg["intensity"][year] - mod["intensity"][year], 3)

    for row in (60, 61, 62):
        copy_row_style(ws, 60, row, 7)
    ws["A60"] = "保守情景 产量(亿吨)"
    ws["A61"] = "适度情景 产量(亿吨)"
    ws["A62"] = "积极情景 产量(亿吨)"
    for col_offset, year in enumerate(KEY_YEARS, start=2):
        ws.cell(60, col_offset).value = cons["production"][year]
        ws.cell(61, col_offset).value = mod["production"][year]
        ws.cell(62, col_offset).value = agg["production"][year]

    for row_idx, tech_idx in zip(range(46, 52), range(6)):
        ws[f"B{row_idx}"] = f"{int(round(cons['shares'][2060][tech_idx]))}%"
        ws[f"C{row_idx}"] = f"{int(round(mod['shares'][2060][tech_idx]))}%"
        ws[f"D{row_idx}"] = f"{int(round(agg['shares'][2060][tech_idx]))}%"

    summary_rows = {55: cons, 56: mod, 57: agg}
    for row, data in summary_rows.items():
        for col_offset, year in enumerate(KEY_YEARS, start=2):
            ws.cell(row, col_offset).value = data["intensity"][year]
    for col_offset, year in enumerate(KEY_YEARS, start=2):
        ws.cell(58, col_offset).value = round(mod["intensity"][year] - cons["intensity"][year], 3)
        ws.cell(59, col_offset).value = round(agg["intensity"][year] - mod["intensity"][year], 3)

    for idx, year in enumerate(YEARS, start=79):
        ws[f"A{idx}"] = year
        ws[f"B{idx}"] = cons["emissions"][year]
        ws[f"C{idx}"] = mod["emissions"][year]
        ws[f"D{idx}"] = agg["emissions"][year]


def update_dashboard_sheet(ws: Worksheet, scenario_paths: Dict[str, Dict[str, Dict[int, float] | Dict[int, List[float]]]]) -> None:
    cons = scenario_paths["conservative"]
    mod = scenario_paths["moderate"]
    agg = scenario_paths["aggressive"]
    ws["E3"] = "适度情景产量(参考)"
    for idx, year in enumerate(YEARS, start=4):
        ws[f"A{idx}"] = year
        ws[f"B{idx}"] = cons["intensity"][year]
        ws[f"C{idx}"] = mod["intensity"][year]
        ws[f"D{idx}"] = agg["intensity"][year]
        ws[f"E{idx}"] = mod["production"][year]
        ws[f"F{idx}"] = cons["emissions"][year]
        ws[f"G{idx}"] = mod["emissions"][year]
        ws[f"H{idx}"] = agg["emissions"][year]


def update_finance_estimate_sheet(ws: Worksheet, scenario_paths: Dict[str, Dict[str, Dict[int, float] | Dict[int, List[float]]]]) -> None:
    cons = scenario_paths["conservative"]
    mod = scenario_paths["moderate"]
    agg = scenario_paths["aggressive"]
    ws["A2"] = "输入行业融资组合参数 → 自动计算碳强度并对比三情景基准（产量列采用适度情景参考路径）"
    ws["D6"] = 0.59
    ws["B12"] = "产量(亿吨,适度参考)"
    for idx, year in enumerate(YEARS, start=13):
        ws[f"A{idx}"] = year
        ws[f"B{idx}"] = mod["production"][year]
        ws[f"C{idx}"] = cons["intensity"][year]
        ws[f"D{idx}"] = cons["emissions"][year]
        ws[f"E{idx}"] = mod["intensity"][year]
        ws[f"F{idx}"] = mod["emissions"][year]
        ws[f"G{idx}"] = agg["intensity"][year]
        ws[f"H{idx}"] = agg["emissions"][year]


def update_finance_enterprise_sheet(ws: Worksheet, scenario_paths: Dict[str, Dict[str, Dict[int, float] | Dict[int, List[float]]]]) -> None:
    conservative_2030 = scenario_paths["conservative"]["intensity"][2030]
    moderate_2030 = scenario_paths["moderate"]["intensity"][2030]
    aggressive_2030 = scenario_paths["aggressive"]["intensity"][2030]

    for row in range(6, 11):
        ws[f"E{row}"] = f'=IF(D{row}="","",D{row}-{conservative_2030:.3f})'
        ws[f"F{row}"] = f'=IF(D{row}="","",D{row}-{moderate_2030:.3f})'
        ws[f"G{row}"] = f'=IF(D{row}="","",D{row}-{aggressive_2030:.3f})'
        ws[f"H{row}"] = (
            f'=IF(D{row}="","",IF(D{row}<={aggressive_2030:.3f},"领跑者",'
            f'IF(D{row}<={moderate_2030:.3f},"积极转型",IF(D{row}<={conservative_2030:.3f},"跟随转型","需重点关注"))))'
        )
        ws[f"J{row}"] = (
            f'=IF(D{row}="","",IF(H{row}="领跑者","维持融资，探索可持续挂钩产品",'
            f'IF(H{row}="积极转型","适度支持转型投资",'
            f'IF(H{row}="跟随转型","增加技改融资引导",'
            f'IF(H{row}="需重点关注","评估退出策略或附加减排条件","")))))'
        )

    ws["B14"] = f'<= {aggressive_2030:.3f} tCO2/t'
    ws["B15"] = f'{aggressive_2030:.3f} - {moderate_2030:.3f}'
    ws["B16"] = f'{moderate_2030:.3f} - {conservative_2030:.3f}'
    ws["B17"] = f'> {conservative_2030:.3f} tCO2/t'
    ws["C14"] = "达到或超过积极情景标准"
    ws["C15"] = "介于积极与适度情景之间"
    ws["C16"] = "介于适度与保守情景之间"
    ws["C17"] = "高于保守情景标准"


def generate_summary(scenario_paths: Dict[str, Dict[str, Dict[int, float] | Dict[int, List[float]]]]) -> str:
    lines = [
        "# 水泥行业转型路径模型校准说明",
        "",
        "## 1. 校准目标",
        "",
        "本次校准按照“自上而下”与“自下而上”相结合的方法，对三种情景的水泥产量、碳强度与技术路径进行重新标定，重点满足 2030/2035 年碳强度节点要求。",
        "",
        "## 2. 三情景关键假设",
        "",
        "| 情景 | 2030产量(亿吨) | 2035产量(亿吨) | 2030碳强度(tCO2/t) | 2035碳强度(tCO2/t) | 2030较2024降幅 | 2035较2024降幅 | 设定逻辑 |",
        "|---|---:|---:|---:|---:|---:|---:|---|",
    ]
    for key in ["conservative", "moderate", "aggressive"]:
        config = SCENARIOS[key]
        data = scenario_paths[key]
        lines.append(
            f"| {config.label} | {data['production'][2030]:.1f} | {data['production'][2035]:.1f} | {data['intensity'][2030]:.3f} | {data['intensity'][2035]:.3f} | {abs(data['drops'][2030]) * 100:.1f}% | {abs(data['drops'][2035]) * 100:.1f}% | {config.scenario_text} |"
        )
    lines.extend(
        [
            "",
            "## 3. 关键节点结果",
            "",
            "| 年份 | 保守情景碳强度 | 适度情景碳强度 | 积极情景碳强度 | 保守CO2(亿吨) | 适度CO2(亿吨) | 积极CO2(亿吨) |",
            "|---|---:|---:|---:|---:|---:|---:|",
        ]
    )
    for year in [2024, 2030, 2035, 2040, 2050, 2060]:
        cons = scenario_paths["conservative"]
        mod = scenario_paths["moderate"]
        agg = scenario_paths["aggressive"]
        lines.append(
            f"| {year} | {cons['intensity'][year]:.3f} | {mod['intensity'][year]:.3f} | {agg['intensity'][year]:.3f} | {cons['emissions'][year]:.2f} | {mod['emissions'][year]:.2f} | {agg['emissions'][year]:.2f} |"
        )
    lines.extend(
        [
            "",
            "## 4. 自上而下 / 自下而上一致性说明",
            "",
            "1. **宏观侧**：三情景采用不同的产量下降路径，保守 > 适度 > 积极，保证总排放下降节奏与情景强弱相匹配。",
            "2. **技术侧**：2030年前的减排主要来自需求收缩、能效提升、替代燃料和熟料替代，CCUS在2035年前仅贡献少量减排。",
            "3. **阶段性加速**：2030-2035 年起，适度与积极情景的 CCUS 和低碳熟料渗透明显提速，支撑中长期更深的碳强度下降。",
            "4. **终局一致性**：三情景均保持 2060 年近零排放/碳中和导向，其中积极情景深度脱碳最强，保守情景则更多依赖后期成熟化部署。",
        ]
    )
    return "\n".join(lines) + "\n"


def update_workbook(input_path: Path, output_path: Path) -> None:
    workbook = load_workbook(input_path)
    scenario_paths = {key: build_paths(config) for key, config in SCENARIOS.items()}

    update_overview_sheet(workbook["模型总览与设计逻辑"], scenario_paths)
    for key, config in SCENARIOS.items():
        update_path_sheet(workbook[config.sheet_name], config, scenario_paths[key])
    update_comparison_sheet(workbook["三情景综合对比"], scenario_paths)
    update_dashboard_sheet(workbook["可视化仪表板"], scenario_paths)
    update_finance_estimate_sheet(workbook["金融机构目标设置（估算）"], scenario_paths)
    update_finance_enterprise_sheet(workbook["金融机构目标设置（企业）"], scenario_paths)

    if input_path.resolve() == output_path.resolve():
        backup_path = input_path.with_suffix(input_path.suffix + ".bak")
        shutil.copy2(input_path, backup_path)

    workbook.save(output_path)
    SUMMARY_PATH.write_text(generate_summary(scenario_paths), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="校准水泥行业转型路径 Excel 模型。")
    parser.add_argument("--input", type=Path, default=WORKBOOK_PATH, help="输入工作簿路径")
    parser.add_argument("--output", type=Path, default=WORKBOOK_PATH, help="输出工作簿路径")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    update_workbook(args.input, args.output)
    print(f"Workbook updated: {args.output}")
    print(f"Summary written: {SUMMARY_PATH}")
